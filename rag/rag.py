# rag.py
# rag本地知识库构建，知识库管理需要rag_store_manager.py文件
from __future__ import annotations

import os
import json
import time
import hashlib
from dataclasses import dataclass, asdict
from typing import Dict, List, Optional, Tuple, Iterable

import numpy as np
import faiss

from utils import settings
from utils.llm import embed_texts, embed_query


# -----------------------------
# Data structures
# -----------------------------
@dataclass
class Chunk:
    vector_id: int
    chunk_id: str
    doc_id: str
    text: str
    source_path: str
    start: int
    end: int


# -----------------------------
# Helpers: IO, hashing, chunking
# -----------------------------
def _now_iso() -> str:
    return time.strftime("%Y-%m-%dT%H:%M:%S", time.localtime())


def _sha256_text(text: str) -> str:
    h = hashlib.sha256()
    h.update(text.encode("utf-8", errors="ignore"))
    return h.hexdigest()


def _norm_path(p: str) -> str:
    return os.path.normpath(os.path.abspath(p))

def _get_short_path_windows(path: str) -> Optional[str]:
    """
    Windows 下尝试将目录路径转为 8.3 short path，避免 faiss 对 Unicode 路径写文件失败。
    若系统未启用 8.3 或转换失败，返回 None。
    """
    if os.name != "nt":
        return None
    try:
        import ctypes  # noqa
        buf = ctypes.create_unicode_buffer(4096)
        r = ctypes.windll.kernel32.GetShortPathNameW(path, buf, 4096)
        if r > 0:
            return buf.value
    except Exception:
        return None
    return None


def _faiss_safe_path(path: str) -> str:
    """
    返回一个尽可能适合 faiss 读写的路径：
    - Windows：把“目录部分”转为 short path，再拼接文件名（文件本身可不存在）
    - 其他系统：原样返回
    """
    ap = os.path.abspath(path)
    if os.name != "nt":
        return ap

    d = os.path.dirname(ap)
    b = os.path.basename(ap)
    short_d = _get_short_path_windows(d)
    if short_d:
        return os.path.join(short_d, b)
    return ap


def _write_text_file(path: str, text: str) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write(text)
        f.flush()
        try:
            os.fsync(f.fileno())
        except Exception:
            # Windows/某些文件系统下 fsync 可能不可用或无必要
            pass


def _write_json_file(path: str, obj: dict) -> None:
    _write_text_file(path, json.dumps(obj, ensure_ascii=False, indent=2))


def _write_jsonl_chunks(path: str, chunks_by_vid: Dict[int, "Chunk"]) -> None:
    lines = []
    for vid in sorted(chunks_by_vid.keys()):
        lines.append(json.dumps(asdict(chunks_by_vid[vid]), ensure_ascii=False))
    _write_text_file(path, "\n".join(lines) + ("\n" if lines else ""))


def read_txt_file(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def read_docx_file(path: str) -> str:
    # 可选依赖：python-docx
    from docx import Document  # type: ignore
    d = Document(path)
    return "\n".join([p.text for p in d.paragraphs if p.text and p.text.strip()])


def read_pdf_file(path: str) -> str:
    """
    优先使用 pypdf；如不可用，会抛出 ImportError。
    """
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception as e:  # pragma: no cover
        raise ImportError("读取 PDF 需要安装 pypdf：pip install pypdf") from e

    reader = PdfReader(path)
    pages = []
    for page in reader.pages:
        try:
            pages.append(page.extract_text() or "")
        except Exception:
            pages.append("")
    return "\n".join([t for t in pages if t.strip()])


def read_xlsx_file(path: str) -> str:
    """
    读取 .xlsx/.xlsm/.xltx/.xltm 为“可检索文本”：
    - 每个 Sheet：找到第一条非空行作为表头
    - 后续每一行：输出为  表头:值  的键值对（用 \\t 分隔）
    - 默认跳过全空行/全空值行
    """
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise ImportError("读取 XLSX 需要安装 openpyxl：pip install openpyxl") from e

    import datetime as _dt

    max_sheets = int(getattr(settings, "XLSX_MAX_SHEETS", 20))
    max_rows = int(getattr(settings, "XLSX_MAX_ROWS_PER_SHEET", 5000))
    max_cols = int(getattr(settings, "XLSX_MAX_COLS_PER_SHEET", 50))

    # 可选：是否把空值也输出为“表头:”
    include_empty = bool(getattr(settings, "XLSX_INCLUDE_EMPTY_VALUES", False))

    def _cell_to_str(v) -> str:
        if v is None:
            return ""
        if isinstance(v, (_dt.datetime, _dt.date)):
            return v.isoformat()
        try:
            return str(v)
        except Exception:
            return ""

    def _normalize_headers(raw_headers: List[str]) -> List[str]:
        """
        - 空表头 -> col_1/col_2...
        - 重复表头 -> name_2/name_3...
        """
        headers: List[str] = []
        seen: Dict[str, int] = {}
        for i, h in enumerate(raw_headers):
            h = (h or "").strip()
            if not h:
                h = f"col_{i+1}"
            cnt = seen.get(h, 0) + 1
            seen[h] = cnt
            if cnt > 1:
                h = f"{h}_{cnt}"
            headers.append(h)
        return headers

    def _extract(data_only: bool) -> str:
        wb = load_workbook(path, read_only=True, data_only=data_only)
        lines: List[str] = []
        sheetnames = list(wb.sheetnames)[:max_sheets]

        for sname in sheetnames:
            ws = wb[sname]
            lines.append(f"# sheet: {sname}")

            headers: Optional[List[str]] = None
            data_rows_written = 0

            for row in ws.iter_rows(values_only=True):
                row = row[:max_cols] if row else []
                row_vals = [_cell_to_str(v) for v in row]

                # 跳过全空行
                if not any(x.strip() for x in row_vals):
                    continue

                # 第一条非空行作为表头
                if headers is None:
                    headers = _normalize_headers(row_vals)
                    continue

                # 行数上限（只统计数据行，不统计表头）
                if data_rows_written >= max_rows:
                    lines.append("... [TRUNCATED: rows limit reached]")
                    break

                # 表头绑定到每个单元格
                pairs: List[str] = []
                non_empty_value_cnt = 0
                for h, v in zip(headers, row_vals):
                    v = (v or "").strip()
                    if v:
                        non_empty_value_cnt += 1
                    if include_empty or v:
                        pairs.append(f"{h}: {v}".rstrip())

                # 默认：整行全空值则跳过
                if not include_empty and non_empty_value_cnt == 0:
                    continue

                lines.append("\t".join(pairs))
                data_rows_written += 1

            lines.append("")  # sheet 分隔空行

        try:
            wb.close()
        except Exception:
            pass

        return "\n".join(lines).strip()

    # 优先取“值”；若几乎读不到内容（例如只有公式但无缓存结果），再读“公式文本”
    text = _extract(data_only=True)
    if not text.strip():
        text = _extract(data_only=False)
    return text

def load_document(path: str) -> Tuple[str, str, str]:
    """
    返回 (doc_id, source_path, text)
    doc_id 由内容 hash 生成；同一内容可稳定复用。
    """
    source_path = _norm_path(path)
    ext = os.path.splitext(path)[1].lower()

    if ext in [".txt", ".md", ".log", ".csv", ".json"]:
        text = read_txt_file(path)
    elif ext in [".docx"]:
        text = read_docx_file(path)
    elif ext in [".pdf"]:
        text = read_pdf_file(path)
    elif ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        text = read_xlsx_file(path)
    else:
        # 兜底：按文本读取
        text = read_txt_file(path)

    text = (text or "").strip()
    doc_id = _sha256_text(text)[:16] if text else _sha256_text(source_path)[:16]
    return doc_id, source_path, text


def chunk_text(text: str, *, chunk_size: int, overlap: int) -> List[Tuple[int, int, str]]:
    """
    简单字符级滑窗切分：返回 [(start, end, chunk_text), ...]
    """
    text = (text or "").strip()
    if not text:
        return []

    if chunk_size <= 0:
        raise ValueError("chunk_size must be > 0")
    if overlap < 0:
        raise ValueError("overlap must be >= 0")
    if overlap >= chunk_size:
        raise ValueError("overlap must be < chunk_size")

    chunks: List[Tuple[int, int, str]] = []
    start = 0
    n = len(text)

    while start < n:
        end = min(start + chunk_size, n)
        ch = text[start:end].strip()
        if ch:
            chunks.append((start, end, ch))
        if end == n:
            break
        start = end - overlap

    return chunks


def chunk_xlsx_rows(text: str) -> List[Tuple[int, int, str]]:
    """
    将 read_xlsx_file 生成的文本按“行”切分为 chunks（每行一个向量）。
    约定：
    - read_xlsx_file 会在每个 sheet 前写入一行：# sheet: <name>
    - 数据行本身是“表头:值\t表头:值...”的键值对形式
    返回：
    - start/end 使用“行号”（从 0 开始）表示，便于追踪；不再代表字符偏移
    """
    text = (text or "").strip()
    if not text:
        return []

    pieces: List[Tuple[int, int, str]] = []
    current_sheet = ""
    row_no = 0

    for raw in text.splitlines():
        line = (raw or "").strip()
        if not line:
            continue

        if line.startswith("# sheet:"):
            current_sheet = line[len("# sheet:"):].strip()
            continue

        # 给每行补充 sheet 上下文，避免跨 sheet 检索时丢失来源
        if current_sheet:
            line_out = f"sheet: {current_sheet}\t{line}"
        else:
            line_out = line

        pieces.append((row_no, row_no, line_out))
        row_no += 1

    return pieces


def _normalize_rows(x: np.ndarray) -> np.ndarray:
    """
    L2 normalize for cosine similarity with inner product.
    """
    if x.ndim != 2:
        raise ValueError("x must be 2D")
    norms = np.linalg.norm(x, axis=1, keepdims=True) + 1e-12
    return x / norms


# -----------------------------
# Faiss RAG store (persistent)
# -----------------------------
class FaissRAG:
    """
    一个可持久化、可增删的轻量 RAG 底座：
    - 使用 IndexIDMap2 + IndexFlatIP（cosine via normalized vectors）
    - 支持 save/load
    - 支持 add_files / remove_doc
    """

    INDEX_FNAME = "index.faiss"
    CHUNKS_FNAME = "chunks.jsonl"
    MANIFEST_FNAME = "manifest.json"

    def __init__(
        self,
        *,
        dim: Optional[int] = None,
        chunk_size: Optional[int] = None,
        overlap: Optional[int] = None,
    ) -> None:
        self.dim: Optional[int] = dim
        self.chunk_size = int(chunk_size if chunk_size is not None else getattr(
            settings, "CHUNK_SIZE", 900))
        self.overlap = int(overlap if overlap is not None else getattr(settings, "CHUNK_OVERLAP", 150))

        # faiss index: created lazily when dim is known
        self.index: Optional[faiss.Index] = None

        # metadata
        self.chunks_by_vid: Dict[int, Chunk] = {}
        self.docs: Dict[str, dict] = {}  # doc_id -> manifest entry
        self.next_vector_id: int = 1

    # --------- state ----------
    def is_empty(self) -> bool:
        return (self.index is None) or (getattr(self.index, "ntotal", 0) == 0)

    def _ensure_index(self, dim: int) -> None:
        if self.index is not None:
            return
        base = faiss.IndexFlatIP(dim)
        self.index = faiss.IndexIDMap2(base)
        self.dim = dim

    # --------- persistence ----------
    @classmethod
    def store_paths(cls, store_dir: str) -> Dict[str, str]:
        d = os.path.abspath(store_dir)
        return {
            "store_dir": d,
            "index": os.path.join(d, cls.INDEX_FNAME),
            "chunks": os.path.join(d, cls.CHUNKS_FNAME),
            "manifest": os.path.join(d, cls.MANIFEST_FNAME),
        }

    def save(self, store_dir: str) -> None:
        p = self.store_paths(store_dir)
        os.makedirs(p["store_dir"], exist_ok=True)

        # 统一使用 tmp 文件，确保写入全成功后再“提交”
        index_final = p["index"]
        chunks_final = p["chunks"]
        manifest_final = p["manifest"]

        index_tmp = index_final + ".tmp"
        chunks_tmp = chunks_final + ".tmp"
        manifest_tmp = manifest_final + ".tmp"

        # 先准备 manifest（但先不落最终文件）
        manifest = {
            "version": 1,
            "saved_at": _now_iso(),
            "dim": self.dim,
            "metric": "cosine_ip",
            "chunk_size": self.chunk_size,
            "overlap": self.overlap,
            "next_vector_id": self.next_vector_id,
            "docs": self.docs,
            "ntotal": int(self.index.ntotal) if self.index is not None else 0,
        }

        # 1) 写 tmp index（最容易失败的步骤先做；失败就直接抛异常，不污染已有库）
        if self.index is not None and getattr(self.index, "ntotal", 0) > 0:
            safe_index_tmp = _faiss_safe_path(index_tmp)
            faiss.write_index(self.index, safe_index_tmp)
        else:
            # 空库：不需要 index 文件
            if os.path.exists(index_tmp):
                os.remove(index_tmp)

        # 2) 写 tmp chunks / tmp manifest（Python 对 Unicode 路径没问题）
        _write_jsonl_chunks(chunks_tmp, self.chunks_by_vid)
        _write_json_file(manifest_tmp, manifest)

        # 3) 提交（原子替换）
        # 3.1 index
        if self.index is not None and getattr(self.index, "ntotal", 0) > 0:
            os.replace(index_tmp, index_final)
        else:
            if os.path.exists(index_final):
                os.remove(index_final)

        # 3.2 chunks / manifest
        os.replace(chunks_tmp, chunks_final)
        os.replace(manifest_tmp, manifest_final)

    @classmethod
    def load(cls, store_dir: str) -> "FaissRAG":
        p = cls.store_paths(store_dir)
        rag = cls(dim=None)

        # no manifest => treat as empty store
        if not os.path.exists(p["manifest"]):
            return rag

        with open(p["manifest"], "r", encoding="utf-8") as f:
            m = json.load(f)

        rag.dim = m.get("dim")
        rag.chunk_size = int(m.get("chunk_size", rag.chunk_size))
        rag.overlap = int(m.get("overlap", rag.overlap))
        rag.next_vector_id = int(m.get("next_vector_id", 1))
        rag.docs = dict(m.get("docs", {}))

        # load chunks
        rag.chunks_by_vid = {}
        if os.path.exists(p["chunks"]):
            with open(p["chunks"], "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    obj = json.loads(line)
                    c = Chunk(**obj)
                    rag.chunks_by_vid[c.vector_id] = c

        # load index (optional)
        if os.path.exists(p["index"]):
            rag.index = faiss.read_index(_faiss_safe_path(p["index"]))
        else:
            # 若 index 不存在，说明库不可检索（通常是写入失败导致的不一致状态）
            # 为避免出现“docs 有但 empty=True”的假象，这里把元信息也视为无效
            rag.index = None
            rag.docs = {}
            rag.chunks_by_vid = {}
            rag.next_vector_id = 1
            return rag

        return rag

    # --------- ingestion ----------
    def _embed_chunks(self, texts: List[str]) -> np.ndarray:
        vecs = embed_texts(texts)  # expect List[List[float]] or np.ndarray
        arr = np.array(vecs, dtype=np.float32)
        if arr.ndim != 2:
            raise ValueError("embed_texts must return a 2D array-like [n, dim]")
        arr = _normalize_rows(arr)
        return arr

    def add_files(self, file_paths: Iterable[str]) -> Dict[str, dict]:
        """
        增量入库：返回新增 doc_id -> entry
        说明：同一 doc_id（内容相同）默认跳过；如要强制重建，请先 remove_doc。
        """
        added: Dict[str, dict] = {}
        for path in file_paths:
            doc_id, source_path, text = load_document(path)
            if not text:
                continue

            # Skip if already present (same content hash)
            if doc_id in self.docs:
                continue
            ext = os.path.splitext(source_path)[1].lower()
            if ext in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
                # Excel：逐行向量化（每行一个向量），不使用滑窗 overlap
                pieces = chunk_xlsx_rows(text)
            else:
                pieces = chunk_text(text, chunk_size=self.chunk_size, overlap=self.overlap)
            if not pieces:
                continue

            # Embed first to infer dim if needed
            chunk_texts = [t for (_, _, t) in pieces]
            vecs = self._embed_chunks(chunk_texts)
            dim = int(vecs.shape[1])
            self._ensure_index(dim)

            # Allocate vector ids
            vids = np.arange(self.next_vector_id, self.next_vector_id + vecs.shape[0], dtype=np.int64)
            self.next_vector_id = int(vids[-1] + 1)

            # Add to faiss
            assert self.index is not None
            self.index.add_with_ids(vecs, vids)

            # Save chunk metadata
            for i, (start, end, ch_text) in enumerate(pieces):
                vid = int(vids[i])
                chunk_id = f"{doc_id}::chunk_{i:06d}" if ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"] else f"{doc_id}::row_{i:06d}"
                self.chunks_by_vid[vid] = Chunk(
                    vector_id=vid,
                    chunk_id=chunk_id,
                    doc_id=doc_id,
                    text=ch_text,
                    source_path=source_path,
                    start=int(start),
                    end=int(end),
                )

            entry = {
                "doc_id": doc_id,
                "source_path": source_path,
                "sha256": _sha256_text(text),
                "n_chunks": int(vecs.shape[0]),
                "vector_ids": [int(v) for v in vids.tolist()],
                "created_at": _now_iso(),
            }
            self.docs[doc_id] = entry
            added[doc_id] = entry

        return added

    # --------- deletion ----------
    def remove_doc(self, *, doc_id: Optional[str] = None, source_path: Optional[str] = None) -> bool:
        """
        删除一个文档对应的所有 chunks。
        返回：是否删除成功（找到了并删除）。
        """
        if doc_id is None and source_path is None:
            raise ValueError("Either doc_id or source_path must be provided")

        if doc_id is None and source_path is not None:
            sp = _norm_path(source_path)
            for did, entry in self.docs.items():
                if _norm_path(entry.get("source_path", "")) == sp:
                    doc_id = did
                    break

        if not doc_id or doc_id not in self.docs:
            return False

        entry = self.docs[doc_id]
        vids = entry.get("vector_ids", [])
        if not vids:
            self.docs.pop(doc_id, None)
            return True

        # Remove from FAISS
        if self.index is not None and getattr(self.index, "ntotal", 0) > 0:
            ids = np.array(vids, dtype=np.int64)
            _ = self.index.remove_ids(ids)

        # Remove metadata
        for vid in vids:
            self.chunks_by_vid.pop(int(vid), None)

        self.docs.pop(doc_id, None)
        return True

    # --------- retrieval ----------
    def search(self, query: str, *, top_k: int = 8) -> List[dict]:
        if self.is_empty():
            return []

        qv = embed_query(query)
        q = np.array([qv], dtype=np.float32)
        if q.ndim != 2:
            raise ValueError("embed_query must return a 1D array-like [dim]")
        q = _normalize_rows(q)

        assert self.index is not None
        scores, ids = self.index.search(q, top_k)
        scores = scores[0].tolist()
        ids = ids[0].tolist()

        hits: List[dict] = []
        for score, vid in zip(scores, ids):
            if vid == -1:
                continue
            c = self.chunks_by_vid.get(int(vid))
            if not c:
                continue
            hits.append(
                {
                    "score": float(score),
                    "vector_id": int(vid),
                    "chunk_id": c.chunk_id,
                    "doc_id": c.doc_id,
                    "source_path": c.source_path,
                    "start": c.start,
                    "end": c.end,
                    "text": c.text,
                }
            )
        return hits

    # --------- inspection ----------
    def list_docs(self) -> List[dict]:
        out = []
        for did, entry in self.docs.items():
            out.append(
                {
                    "doc_id": did,
                    "source_path": entry.get("source_path"),
                    "n_chunks": entry.get("n_chunks", 0),
                    "created_at": entry.get("created_at"),
                }
            )
        return sorted(out, key=lambda x: (x.get("created_at") or "", x["doc_id"]))

# ---------------------------
# 本地快速测试
# ---------------------------
if __name__ == "__main__":
    # 简单自测：增量入库 + 检索
    store_dir = getattr(settings, "RAG_STORE_DIR", "rag_store")
    rag = FaissRAG.load(store_dir)

    test_file = getattr(settings, "RAG_SMOKE_TEST_FILE", None)
    if test_file and os.path.exists(test_file):
        print(f"[ingest] {test_file}")
        rag.add_files([test_file])
        rag.save(store_dir)

    q = "这份材料中有哪些关于价格战、利润下滑的表述？请给出原文证据。"
    hits = rag.search(q, top_k=5)

    print("\n=== Top Hits ===")
    for i, h in enumerate(hits, 1):
        print(f"\n[{i}] score={h['score']:.4f}  {h['chunk_id']}  ({os.path.basename(h['source_path'])})")
        print(h["text"][:300])
